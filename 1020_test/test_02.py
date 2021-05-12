import requests
import openpyxl as px
from bs4 import BeautifulSoup
from openpyxl import Workbook

# URLの取得と割当
"""
load_url = "https://www.buyma.com/buyer/6020144.html"
html = requests.get(load_url)
soup = BeautifulSoup(html.content, "html.parser")
"""
file_path = 'text.xlsx'
wb = px.Workbook()
ws = wb.active


base_url1 = "https://www.buyma.com/buyer/"
base_url2 = ".html"
num = 1
base_urlm = ""
while num <= 9999999:
    if len(str(num)) == 1:
        base_urlm = "000000"
    if len(str(num)) == 2:
        base_urlm = "00000"
    if len(str(num)) == 3:
        base_urlm = "0000"
    if len(str(num)) == 4:
        base_urlm = "000"
    if len(str(num)) == 5:
        base_urlm = "00"
    if len(str(num)) == 6:
        base_urlm = "0"
    if len(str(num)) == 7:
        base_urlm = ""
    load_url = base_url1 + base_urlm + str(num) + base_url2
    html = requests.get(load_url)
    soup = BeautifulSoup(html.content, "html.parser")

    ws.cell(row=num+1, column=1, value=str(load_url))
 #   print(str(load_url))
    num +=1

# sheet.cell(”行", "列" ,"セルへ入力したい値")

"""
# ショップ名 shop_name
buyer_name = soup.find(id="buyer_name")
for element10 in buyer_name.find_all("a"):
    t = element10.text
    shop_name = t

# 評価 shop_eva
for element20 in soup.find_all(class_="buyer_eva_total_count"):
    t = element20.text
    shop_eva = t

# 開始日 shop_start
profile = soup.find(class_="profile_txt")
for element50 in profile.select("dd", limit=1):
    t = element50.text
    shop_start = t.strip()

# テスト的に出力
print(shop_name + '\n\n' + shop_eva + '\n\n' +  shop_start)

"""
wb.save('openpyxl.xlsx')
