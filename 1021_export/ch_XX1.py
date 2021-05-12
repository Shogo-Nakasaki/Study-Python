import requests
import openpyxl as px
from bs4 import BeautifulSoup
from openpyxl import Workbook

# エクセル導入
file_path = 'text.xlsx'
wb = px.Workbook()
ws = wb.active

#変数
ta = 7610000       #開始地点
tb = 7620000    #終了地点

# URLの作成と導入
base_url1 = "https://www.buyma.com/buyer/"
base_url2 = ".html"
num = ta
i=0
base_urlm = ""
while num <= tb:
    if len(str(num)) == 1:
        base_urlm = "000000"
    if len(str(num)) == 2:
        base_urlm = "00000"
    if len(str(num)) == 3:
        base_urlm = "0000"
    if len(str(num)) == 4:
        base_urlm = "000"
    if len(str(num)) == 5:
        base_urlm = "00"
    if len(str(num)) == 6:
        base_urlm = "0"
    if len(str(num)) == 7:
        base_urlm = ""
    load_url = base_url1 + base_urlm + str(num) + base_url2
    html = requests.get(load_url)
    soup = BeautifulSoup(html.content, "html.parser")

    # 無題を省く
    check = soup.find("title")
    print(str(load_url))
    # print(str(check))
    num +=1
    if "バイマ" in str(check):
        # print("ユーザーなし")
        continue
    if "ログイン" in str(check):
        # print("ユーザーなし")
        continue
    
    # 開始日 shop_start
    c = 0
    profile = soup.find(class_="profile_txt")
    for element50 in profile.select("dd", limit=1):
        t = element50.text
        shop_start = t.strip()
        # 2019年と2020年以外は除外
        if "2019" in str(shop_start):
            c+=1
            # print("2019年だ")
            continue
        elif "2020" in str(shop_start):
            c+=1
            # print("2020年だ")
            continue
    if c==0:
        # print("2019年でも2020年でもない")
        continue

    # ショップの種類 shop_label
    buyer_label = soup.find(class_="label")
    t = buyer_label.text
    shop_label = t
    if "PERSONAL SHOPPER" != t:
        continue

    # ショップ名 shop_name
    buyer_name = soup.find(id="buyer_name")
    for element10 in buyer_name.find_all("a"):
        t = element10.text
        shop_name = t

    # 評価 shop_eva
    for element20 in soup.find_all(class_="buyer_eva_total_count"):
        t = element20.text
        shop_eva = t

    i +=1
    ws.cell(row=i  , column=1, value=str(load_url))
    ws.cell(row=i  , column=2, value=str(shop_name))
    ws.cell(row=i  , column=3, value=str(shop_eva))
    ws.cell(row=i  , column=4, value=str(shop_start))
    ws.cell(row=i  , column=5, value=str(shop_label))
    wb.save('wx761_1024.xlsx')
