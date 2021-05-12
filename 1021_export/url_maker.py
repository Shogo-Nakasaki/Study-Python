# import requests
import openpyxl as px
# from bs4 import BeautifulSoup
from openpyxl import Workbook

# エクセル導入
file_path = 'text.xlsx'
wb = px.Workbook()
ws = wb.active

# URLの作成と導入
base_url1 = "https://www.buyma.com/buyer/"
base_url2 = ".html"
num = 1000000
a=0
base_urlm = ""
while num <= 10000000:
#while num <= 1001000:
    if len(str(num)) == 1:
        base_urlm = "000000"
    if len(str(num)) == 2:
        base_urlm = "00000"
    if len(str(num)) == 3:
        base_urlm = "0000"
    if len(str(num)) == 4:
        base_urlm = "000"
    if len(str(num)) == 5:
        base_urlm = "00"
    if len(str(num)) == 6:
        base_urlm = "0"
    if len(str(num)) == 7:
        base_urlm = ""
    load_url = base_url1 + base_urlm + str(num) + base_url2
    #html = requests.get(load_url)
    #soup = BeautifulSoup(html.content, "html.parser")
    print(str(load_url))
    num+=1
    a+=1
    if a<=1000000:
        ws.cell(row=a  , column=1, value=str(load_url))
    elif a<=2000000:
        a=a-1000000
        ws.cell(row=a  , column=2, value=str(load_url))
        a=a+1000000
    elif a<=3000000:
        a=a-2000000
        ws.cell(row=a  , column=3, value=str(load_url))
        a=a+2000000
    elif a<=4000000:
        a=a-3000000
        ws.cell(row=a  , column=4, value=str(load_url))
        a=a+3000000
    elif a<=5000000:
        a=a-4000000
        ws.cell(row=a  , column=5, value=str(load_url))
        a=a+4000000
    elif a<=6000000:
        a=a-5000000
        ws.cell(row=a  , column=6, value=str(load_url))
        a=a+5000000
    elif a<=7000000:
        a=a-6000000
        ws.cell(row=a  , column=7, value=str(load_url))
        a=a+6000000
    elif a<=8000000:
        a=a-7000000
        ws.cell(row=a  , column=8, value=str(load_url))
        a=a+7000000
    elif a<=9000000:
        a=a-8000000
        ws.cell(row=a  , column=9, value=str(load_url))
        a=a+8000000
    elif a<=10000000:
        a=a-9000000
        ws.cell(row=a  , column=10, value=str(load_url))
        a=a+9000000
    
wb.save('url_1021.xlsx')
