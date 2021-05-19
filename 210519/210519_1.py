# 英語のテストを作るためにExcelに表入力するプログラム
import openpyxl as px
from openpyxl import Workbook
import openpyxl

wb = openpyxl.Workbook()
wb.save("System英単語.xlsx")
ws = wb.active
num = 0
print("終了するには3を入力してください\n\n")
while(True):
    # テスト内容の入力
    text1 = input(f"英単語>>")
    if text1 == '3':
        print("作成を終了します")
        break
    text2 = input(f"日本語>>")
    num += 1

    # Excelに自動入力
    ws.cell(row=num, column=1, value=str(num))
    ws.cell(row=num, column=2, value=str(text1))
    ws.cell(row=num, column=3, value=str(text2))
    wb.save("System英単語.xlsx")
# 終了のため、再度保存
wb.save("System英単語.xlsx")